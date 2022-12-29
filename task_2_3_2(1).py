import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit


class Vacancy:
    """
    Класс для представления вакансии.
    Attribytes:
        vacancy ({}): Словарь с данными о вакансии
    """
    currency_rate = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        """
        Инициализирует объект Vacancy, выполняет конвертацию для целочисленных полей
        Args:
            vacancy ({}): словарь с данными
        """
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency_rate[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


class DataSet:
    """
        Отвечает за чтение и подготовку данных из CSV-файла (универсальный парсер CSV)
    """

    def __init__(self, file_name, vacancy_name):
        """
        Инициализирует объект Vacancy, выполняет конвертацию для целочисленных полей
        Args:
            file_name (str) название файла
            vacancy_name (str) название вакансии
        """
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    @staticmethod
    def increment_dict(dictionary, key, amount):
        """Увеличивает словарь на определённое количество"""
        if key in dictionary:
            dictionary[key] += amount
        else:
            dictionary[key] = amount

    @staticmethod
    def get_average(dictionary):
        """Находит среднее значение словаря"""
        new_dictionary = {}
        for key, values in dictionary.items():
            new_dictionary[key] = int(sum(values) / len(values))
        return new_dictionary

    def csv_reader(self):
        """Читает csv файл"""
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    def get_data(self):
        """Получает данные о вакансиях на освновании полей созданного объекта

        :returns (data1, vacancies_number, data2, vacancies_number_by_name, data3, data5): Статистика по зп, статистика по
                числу вакансий, статистика вакансий по ЗП, статистика вакансий по названию, статистика вакансий по городам
        """
        salary = {}
        salary_of_vacancy_name = {}
        salary_city = {}
        count_of_vacancies = 0

        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            self.increment_dict(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment_dict(salary_of_vacancy_name, vacancy.year, [vacancy.salary_average])
            self.increment_dict(salary_city, vacancy.area_name, [vacancy.salary_average])
            count_of_vacancies += 1

        vacancies_number = dict([(key, len(value)) for key, value in salary.items()])
        vacancies_number_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])

        if not salary_of_vacancy_name:
            salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
            vacancies_number_by_name = dict([(key, 0) for key, value in vacancies_number.items()])

        data1 = self.get_average(salary)
        data2 = self.get_average(salary_of_vacancy_name)
        data3 = self.get_average(salary_city)

        data4 = {}
        for year, salaries in salary_city.items():
            data4[year] = round(len(salaries) / count_of_vacancies, 4)
        data4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in data4.items()]))
        data4.sort(key=lambda a: a[-1], reverse=True)
        data5 = data4.copy()
        data4 = dict(data4)
        data3 = list(filter(lambda a: a[0] in list(data4.keys()), [(key, value) for key, value in data3.items()]))
        data3.sort(key=lambda a: a[-1], reverse=True)
        data3 = dict(data3[:10])
        data5 = dict(data5[:10])

        return data1, vacancies_number, data2, vacancies_number_by_name, data3, data5

    @staticmethod
    def print_data(data1, data2, data3, data4, data5, data6):
        """Печатает данные в консоль"""
        print('Динамика уровня зарплат по годам: {0}'.format(data1))
        print('Динамика количества вакансий по годам: {0}'.format(data2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(data3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(data4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(data5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(data6))


class InputConnect:
    """отвечает за обработку параметров вводимых пользователем:
    фильтры, сортировка, диапазон вывода, требуемые столбцы, а также за печать таблицы на экран
    """

    def __init__(self):
        """Создаёт необходимые файлы и печатает на экран в зависимости от пользовательского ввода"""
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        data1, data2, data3, data4, data5, data6 = dataset.get_data()
        dataset.print_data(data1, data2, data3, data4, data5, data6)

        report = Report(self.vacancy_name, data1, data2, data3, data4, data5, data6)
        report.create_excel_table()
        report.save('report.xlsx')
        report.create_image()
        report.create_pdf()


class Report:
    """Основной класс с логикой. Генерирует отчеты XLSX,  PNG, PDF"""

    def __init__(self, vacancy_name, data1, data2, data3, data4, data5, data6):
        """Инициализирует объект отчёта"""
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.data1 = data1
        self.data2 = data2
        self.data3 = data3
        self.data4 = data4
        self.data5 = data5
        self.data6 = data6

    def create_excel_table(self):
        """Создаёт табличный файл отчёта"""
        ws_active = self.wb.active
        ws_active.title = 'Статистика по годам'
        ws_active.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий',
                          'Количество вакансий - ' + self.vacancy_name])
        for year in self.data1.keys():
            ws_active.append([year, self.data1[year], self.data3[year], self.data2[year], self.data4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws_active.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.data5.items(), self.data6.items()):
            data.append([city1, value1, '', city2, value2])
        ws_sheet = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws_sheet.append(row)

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws_sheet.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            ws_active[col + '1'].font = font_bold
            ws_sheet[col + '1'].font = font_bold

        for index, _ in enumerate(self.data5):
            ws_sheet['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                ws_sheet[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for row, _ in enumerate(self.data1):
            for col in 'ABCDE':
                ws_active[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def create_image(self):
        """Создаёт PNG-ищображение графиков"""
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        bar1 = ax1.bar(np.array(list(self.data1.keys())) - 0.4, self.data1.values(), width=0.4)
        bar2 = ax1.bar(np.array(list(self.data1.keys())), self.data3.values(), width=0.4)
        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax1.grid(axis='y')
        ax1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        ax1.set_xticks(np.array(list(self.data1.keys())) - 0.2, list(self.data1.keys()), rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)

        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = ax2.bar(np.array(list(self.data2.keys())) - 0.4, self.data2.values(), width=0.4)
        bar2 = ax2.bar(np.array(list(self.data2.keys())), self.data4.values(), width=0.4)
        ax2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
                   prop={'size': 8})
        ax2.set_xticks(np.array(list(self.data2.keys())) - 0.2, list(self.data2.keys()), rotation=90)
        ax2.grid(axis='y')
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)

        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.data5.keys()))]),
                 list(reversed(list(self.data5.values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')

        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.data6.values()])
        ax4.pie(list(self.data6.values()) + [other], labels=list(self.data6.keys()) + ['Другие'],
                textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')

    def create_pdf(self):
        """Создаёт PDF-документ со статистикой"""
        env = Environment(loader=FileSystemLoader('../templates'))
        template = env.get_template("pdf.html")
        data = []
        for year in self.data1.keys():
            data.append([year, self.data1[year], self.data2[year], self.data3[year], self.data4[year]])

        for key in self.data6:
            self.data6[key] = round(self.data6[key] * 100, 2)

        pdf_template = template.render(
            {
                'name': self.vacancy_name,
                'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
                'data': data,
                'data5': self.data5,
                'data6': self.data6
            }
        )

        config = pdfkit.configuration(wkhtmltopdf=r'C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe')
        # pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

    def save(self, filename):
        """Сохраняет xlsx файл"""
        self.wb.save(filename=filename)